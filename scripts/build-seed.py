#!/usr/bin/env python3
"""
build-seed.py — read strong-curves.xlsx, emit src/seed/strongCurves.ts.

Source of truth for the two built-in routines:

  * Strong Curves — Bootyful Beginnings (8 weeks)
  * Strong Curves — Bodyweight (4 weeks)

The xlsx has one sheet per week ("Bootyful Beginnings Week 1", etc.).
Each sheet packs 4 workout days. Workout content lives in column 12,
laid out as:

    DAY N - WORKOUT X     <- header
    A1: <exercise>        <- row r
    3 sets, 10-20 reps    <- row r + 1
    <optional notes>      <- row r + 2 (sometimes)
    A2: <exercise>
    ...

Exercises prefixed `A1:`, `A2:`, `B1:`, `B2:` form supersets; the rest
are plain (single-exercise) blocks. After the last exercise the column
holds image-anchor labels — we stop reading once we see a name without
a sets/reps spec on the next row.

Muscle tags, measurement types, default rest, and the `usesBarbell`
flag aren't in the xlsx. We carry them in EXERCISE_OVERRIDES below;
unknown exercises fall through to safe defaults and are listed at the
end of the run for human review.

Run via:

    pnpm seed:build
"""

from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install it with `python -m pip install openpyxl`.")

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "strong-curves.xlsx"
OUTPUT_PATH = ROOT / "src" / "seed" / "strongCurves.ts"

WORKOUT_COL = 12  # 1-indexed column where the workout content lives.

# Workout pattern within a week sheet:
#   Day 1 -> Workout A
#   Day 2 -> Workout B
#   Day 3 -> rest (omitted from sheet)
#   Day 4 -> Workout A (mirrors day 1; treated as its own session in-app)
#   Day 5 -> Workout C
#   Days 6, 7 -> rest
DAY_PATTERN = re.compile(r"DAY\s+(\d+)\s*-\s*WORKOUT\s+([A-Z])", re.IGNORECASE)
SUPERSET_PATTERN = re.compile(r"^([A-Z])(\d+)\s*:\s*(.+)$")
SETS_REPS_PATTERN = re.compile(
    r"^(?P<sets>\d+)\s+sets?\s*[,:]?\s*"
    r"(?P<spec>.*?)\s*$",
    re.IGNORECASE,
)
TIME_PATTERN = re.compile(r"\(?(\d+)\s*-\s*(\d+)\s*seconds?\)?", re.IGNORECASE)
REPS_PATTERN = re.compile(r"(\d+)\s*-\s*(\d+)\s*reps?", re.IGNORECASE)
SINGLE_REPS_PATTERN = re.compile(r"^(\d+)\s*reps?\b", re.IGNORECASE)
EACH_SIDE_PATTERN = re.compile(r"each(\s+side)?", re.IGNORECASE)


# --- Seed data classes ------------------------------------------------------

@dataclass(frozen=True)
class SeedExercise:
    id: str
    name: str
    category: str
    primary_muscles: tuple[str, ...]
    secondary_muscles: tuple[str, ...]
    measurement_type: str
    default_rest_seconds: int
    per_side: bool
    uses_barbell: bool


@dataclass
class SeedPlannedExercise:
    exercise_id: str
    set_count: int
    reps_min: int | None = None
    reps_max: int | None = None
    duration_min: int | None = None
    duration_max: int | None = None
    per_side: bool = False
    notes: str | None = None


@dataclass
class SeedBlock:
    type: str  # 'single' | 'superset'
    exercises: list[SeedPlannedExercise] = field(default_factory=list)


@dataclass
class SeedDay:
    day_number: int
    kind: str  # 'workout' | 'rest'
    workout_label: str | None
    blocks: list[SeedBlock] = field(default_factory=list)


@dataclass
class SeedWeek:
    week_number: int
    days: list[SeedDay] = field(default_factory=list)


@dataclass
class SeedRoutine:
    id: str
    name: str
    description: str
    weeks: list[SeedWeek] = field(default_factory=list)


# --- Manual exercise metadata -----------------------------------------------
#
# Lookup keyed by slug (normalised exercise name). Values override the
# defaults applied to unknown exercises. Keep this short and focused —
# exotic variants can fall through to the heuristic.

EXERCISE_OVERRIDES: dict[str, dict] = {
    # --- Glutes -----------------------------------------------------------
    "bodyweight-barbell-glute-bridge": dict(
        category="glute",
        primary=("glutes",),
        secondary=("hamstrings",),
        measurement="weight_reps",
        rest=90,
        uses_barbell=True,
    ),
    "single-leg-glute-bridge": dict(
        category="glute",
        primary=("glutes",),
        secondary=("hamstrings",),
        measurement="reps_each_side",
        rest=60,
        per_side=True,
    ),
    "bodyweight-hip-thrust": dict(
        category="glute",
        primary=("glutes",),
        secondary=("hamstrings",),
        measurement="bodyweight_reps",
        rest=90,
    ),
    "barbell-hip-thrust": dict(
        category="glute",
        primary=("glutes",),
        secondary=("hamstrings",),
        measurement="weight_reps",
        rest=120,
        uses_barbell=True,
    ),
    # --- Quads / hinges --------------------------------------------------
    "bodyweight-barbell-box-squat": dict(
        category="quad",
        primary=("quads", "glutes"),
        secondary=("hamstrings",),
        measurement="weight_reps",
        rest=120,
        uses_barbell=True,
    ),
    "bodyweight-step-up-reverse-lunge-combo": dict(
        category="quad",
        primary=("quads", "glutes"),
        secondary=("hamstrings",),
        measurement="bodyweight_reps",
        rest=60,
        per_side=True,
    ),
    "dumbbell-romanian-deadlift": dict(
        category="hip-hinge",
        primary=("hamstrings", "glutes"),
        secondary=("back",),
        measurement="weight_reps",
        rest=90,
    ),
    "barbell-romanian-deadlift": dict(
        category="hip-hinge",
        primary=("hamstrings", "glutes"),
        secondary=("back",),
        measurement="weight_reps",
        rest=120,
        uses_barbell=True,
    ),
    "bodyweight-hip-hinge-with-dowel": dict(
        category="hip-hinge",
        primary=("hamstrings", "glutes"),
        secondary=("back",),
        measurement="bodyweight_reps",
        rest=60,
    ),
    # --- Push -------------------------------------------------------------
    "barbell-dumbbell-bench-press": dict(
        category="push",
        primary=("chest",),
        secondary=("triceps", "shoulders"),
        measurement="weight_reps",
        rest=90,
        uses_barbell=True,
    ),
    "barbell-bench-press": dict(
        category="push",
        primary=("chest",),
        secondary=("triceps", "shoulders"),
        measurement="weight_reps",
        rest=120,
        uses_barbell=True,
    ),
    "torso-elevated-push-up": dict(
        category="push",
        primary=("chest",),
        secondary=("triceps", "shoulders", "core"),
        measurement="bodyweight_reps",
        rest=60,
    ),
    "dumbbell-military-press": dict(
        category="push",
        primary=("shoulders",),
        secondary=("triceps",),
        measurement="weight_reps",
        rest=90,
    ),
    # --- Pull -------------------------------------------------------------
    "one-arm-dumbbell-row": dict(
        category="pull",
        primary=("back", "lats"),
        secondary=("biceps",),
        measurement="reps_each_side",
        rest=60,
        per_side=True,
    ),
    "front-lat-pulldowns": dict(
        category="pull",
        primary=("lats", "back"),
        secondary=("biceps",),
        measurement="weight_reps",
        rest=90,
    ),
    "towel-row": dict(
        category="pull",
        primary=("back", "lats"),
        secondary=("biceps",),
        measurement="bodyweight_reps",
        rest=60,
    ),
    "negative-chin-up-underhand-grip-pull-down": dict(
        category="pull",
        primary=("lats", "back"),
        secondary=("biceps",),
        measurement="bodyweight_reps",
        rest=90,
    ),
    "standing-single-arm-cable-row": dict(
        category="pull",
        primary=("back", "lats"),
        secondary=("biceps",),
        measurement="reps_each_side",
        rest=60,
        per_side=True,
    ),
    # --- Accessory --------------------------------------------------------
    "bodyweight-dumbbell-step-up": dict(
        category="quad",
        primary=("quads", "glutes"),
        secondary=("hamstrings",),
        measurement="reps_each_side",
        rest=60,
        per_side=True,
    ),
    "bodyweight-or-45-lb-back-extension": dict(
        category="hip-hinge",
        primary=("glutes", "hamstrings"),
        secondary=("back",),
        measurement="weight_reps",
        rest=60,
    ),
    "side-lying-abductions": dict(
        category="accessory",
        primary=("abductors", "glutes"),
        secondary=(),
        measurement="reps_each_side",
        rest=45,
        per_side=True,
    ),
    "side-lying-hip-abduction": dict(
        category="activation",
        primary=("abductors", "glutes"),
        secondary=(),
        measurement="reps_each_side",
        rest=30,
        per_side=True,
    ),
    # --- Core -------------------------------------------------------------
    "rkc-plank": dict(
        category="core",
        primary=("core",),
        secondary=(),
        measurement="time_seconds",
        rest=45,
    ),
    "elevated-foot-rkc-plank": dict(
        category="core",
        primary=("core",),
        secondary=("glutes",),
        measurement="time_seconds",
        rest=45,
    ),
    "side-plank": dict(
        category="core",
        primary=("core",),
        secondary=("abductors",),
        measurement="time_seconds",
        rest=30,
        per_side=True,
    ),
    "side-plank-on-feet": dict(
        category="core",
        primary=("core",),
        secondary=("abductors",),
        measurement="time_seconds",
        rest=30,
        per_side=True,
    ),
}


# --- Helpers ----------------------------------------------------------------

def slugify(name: str) -> str:
    """Lower-case, hyphen-separated, alphanumerics only."""
    s = re.sub(r"[\s/\\–—_]+", "-", name.strip().lower())
    s = re.sub(r"[^a-z0-9-]", "", s)
    s = re.sub(r"-+", "-", s)
    return s.strip("-")


def looks_like_sets_reps(value: str | None) -> bool:
    if not value:
        return False
    v = str(value).strip().lower()
    return bool(re.match(r"^\d+\s+sets?\b", v))


def parse_sets_reps(spec: str) -> tuple[int, dict]:
    """Return (set_count, kwargs) from a spec like '3 sets, 10-20 reps' or
    '1 set (20-120 seconds)'."""
    m = re.match(r"^(\d+)\s+sets?\b", spec, re.IGNORECASE)
    set_count = int(m.group(1)) if m else 1

    kwargs: dict = {}
    if t := TIME_PATTERN.search(spec):
        kwargs["duration_min"] = int(t.group(1))
        kwargs["duration_max"] = int(t.group(2))
    elif r := REPS_PATTERN.search(spec):
        kwargs["reps_min"] = int(r.group(1))
        kwargs["reps_max"] = int(r.group(2))
    elif sr := SINGLE_REPS_PATTERN.search(spec):
        kwargs["reps_min"] = int(sr.group(1))
        kwargs["reps_max"] = int(sr.group(1))

    if EACH_SIDE_PATTERN.search(spec):
        kwargs["per_side"] = True

    return set_count, kwargs


def normalise_exercise_name(raw: str) -> tuple[str, str | None]:
    """Strip A1:/B2:/etc. prefixes; collapse whitespace. Returns
    (clean_name, superset_marker)."""
    cleaned = raw.replace("\n", " ").replace("\r", " ").strip()
    m = SUPERSET_PATTERN.match(cleaned)
    marker = None
    if m:
        marker = m.group(1) + m.group(2)  # e.g. "A1"
        cleaned = m.group(3).strip()
    cleaned = re.sub(r"\s{2,}", " ", cleaned)
    return cleaned, marker


# --- Parsing core -----------------------------------------------------------

def iter_day_headers(ws) -> Iterable[tuple[int, int, str]]:
    """Yield (row, day_number, workout_letter) for each day section."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, WORKOUT_COL).value
        if not isinstance(v, str):
            continue
        m = DAY_PATTERN.search(v)
        if m:
            yield r, int(m.group(1)), m.group(2).upper()


def parse_day(ws, start_row: int, end_row: int) -> list[SeedBlock]:
    """Walk the workout column between header (exclusive) and the next
    header row (exclusive), assembling blocks from (name, sets/reps) pairs.
    Stops at the first non-pair entry — that's the start of the
    image-anchor section below."""
    blocks: list[SeedBlock] = []
    pending: dict[str, SeedBlock] = {}  # superset letter -> block

    r = start_row + 1
    while r < end_row:
        name_cell = ws.cell(r, WORKOUT_COL).value
        if not isinstance(name_cell, str) or not name_cell.strip():
            r += 1
            continue

        # If the next non-empty row isn't a sets/reps spec, we've hit the
        # image-anchor zone — stop.
        next_row = r + 1
        while next_row < end_row and (
            ws.cell(next_row, WORKOUT_COL).value is None
            or not str(ws.cell(next_row, WORKOUT_COL).value).strip()
        ):
            next_row += 1
        if next_row >= end_row:
            break

        spec = ws.cell(next_row, WORKOUT_COL).value
        if not looks_like_sets_reps(spec):
            break

        clean_name, marker = normalise_exercise_name(name_cell)
        set_count, kwargs = parse_sets_reps(str(spec))

        # Optional notes line — a non-empty row after spec that isn't another
        # exercise pair start.
        notes = None
        notes_row = next_row + 1
        if notes_row < end_row:
            notes_value = ws.cell(notes_row, WORKOUT_COL).value
            if isinstance(notes_value, str) and notes_value.strip():
                following = ws.cell(notes_row + 1, WORKOUT_COL).value if notes_row + 1 < end_row else None
                if not looks_like_sets_reps(following):
                    notes = notes_value.strip()
                    next_row = notes_row

        planned = SeedPlannedExercise(
            exercise_id=slugify(clean_name),
            set_count=set_count,
            notes=notes,
            **kwargs,
        )

        if marker:
            letter = marker[0]
            if letter in pending:
                pending[letter].exercises.append(planned)
            else:
                block = SeedBlock(type="superset", exercises=[planned])
                blocks.append(block)
                pending[letter] = block
        else:
            blocks.append(SeedBlock(type="single", exercises=[planned]))
            pending.clear()  # singles break any open superset grouping

        r = next_row + 1

    return blocks


WORKOUT_DAY_SLOTS = [1, 2, 4, 5]  # Days 3, 6, 7 are rest. See SCOPE.md.


def parse_week_sheet(ws, week_number: int) -> SeedWeek:
    """Parse a week sheet positionally — the first workout block becomes
    Day 1, the second Day 2, the third Day 4, the fourth Day 5. The
    source xlsx has a labelling bug from Week 5 onwards (the third
    section is mis-labelled "DAY 1" instead of "DAY 4"); we ignore the
    header's day number and trust the position. Workout label letters
    (A/B/A/C) are taken from the header — those are correct."""
    headers = list(iter_day_headers(ws))
    if not headers:
        return SeedWeek(week_number=week_number, days=[])
    headers.append((ws.max_row + 1, -1, ""))  # sentinel

    days: dict[int, SeedDay] = {}
    for i, (row, _day_num, letter) in enumerate(headers[:-1]):
        if i >= len(WORKOUT_DAY_SLOTS):
            print(
                f"  warning: extra workout section in week {week_number} "
                f"(slot {i+1}) — ignoring",
                file=sys.stderr,
            )
            break
        slot = WORKOUT_DAY_SLOTS[i]
        end_row = headers[i + 1][0]
        blocks = parse_day(ws, row, end_row)
        days[slot] = SeedDay(
            day_number=slot,
            kind="workout",
            workout_label=letter,
            blocks=blocks,
        )

    full_days: list[SeedDay] = []
    for n in range(1, 8):
        if n in days:
            full_days.append(days[n])
        else:
            full_days.append(SeedDay(day_number=n, kind="rest", workout_label=None))
    return SeedWeek(week_number=week_number, days=full_days)


def parse_routine(wb, name: str, description: str, week_sheet_pattern: str, week_count: int) -> SeedRoutine:
    routine = SeedRoutine(id=slugify(name), name=name, description=description)
    for n in range(1, week_count + 1):
        sheet_name = week_sheet_pattern.format(n=n)
        if sheet_name not in wb.sheetnames:
            print(f"  warning: missing sheet {sheet_name!r}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        routine.weeks.append(parse_week_sheet(ws, n))
    return routine


def collect_exercises(routines: Iterable[SeedRoutine]) -> list[SeedExercise]:
    seen: dict[str, SeedExercise] = {}
    name_by_slug: dict[str, str] = {}
    unknown: list[str] = []

    for routine in routines:
        for week in routine.weeks:
            for day in week.days:
                for block in day.blocks:
                    for ex in block.exercises:
                        slug = ex.exercise_id
                        if slug in seen:
                            continue
                        # Look up display name from the latest plan we saw —
                        # we lost it during slugify; reconstruct from the
                        # last `pretty` map. We have no name on the planned
                        # exercise, so we re-derive from the slug below if
                        # needed.
                        # (We recover the original name in collect_exercise_names.)

    # Re-walk to record canonical display names.
    name_by_slug = collect_exercise_names(routines)

    for slug, name in sorted(name_by_slug.items()):
        override = EXERCISE_OVERRIDES.get(slug)
        if override:
            seen[slug] = SeedExercise(
                id=slug,
                name=name,
                category=override.get("category", "other"),
                primary_muscles=tuple(override.get("primary", ())),
                secondary_muscles=tuple(override.get("secondary", ())),
                measurement_type=override.get("measurement", "weight_reps"),
                default_rest_seconds=override.get("rest", 90),
                per_side=override.get("per_side", False),
                uses_barbell=override.get("uses_barbell", False),
            )
        else:
            unknown.append(name)
            seen[slug] = SeedExercise(
                id=slug,
                name=name,
                category="other",
                primary_muscles=(),
                secondary_muscles=(),
                measurement_type="weight_reps",
                default_rest_seconds=60,
                per_side=False,
                uses_barbell=False,
            )

    if unknown:
        print(
            "  note: " + str(len(unknown)) + " exercises lack manual metadata "
            "(falling back to defaults). Add overrides in build-seed.py to "
            "improve muscle tags / rest / barbell flags:",
            file=sys.stderr,
        )
        for name in unknown:
            print(f"    - {name}", file=sys.stderr)

    return list(seen.values())


def collect_exercise_names(routines: Iterable[SeedRoutine]) -> dict[str, str]:
    """We slugified during parsing, losing the display name. Re-walk the
    workbook columns to recover the canonical, full exercise names."""
    # Instead of re-parsing, we attach the display name on a side channel
    # at parse time. Stub for symmetry; real population happens via
    # ROUTINE_NAME_BUFFER.
    return dict(NAME_BY_SLUG)


# --- TS code-gen ------------------------------------------------------------

def ts_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def ts_emit(
    routines: list[SeedRoutine], exercises: list[SeedExercise]
) -> str:
    out: list[str] = []
    out.append("// AUTOGENERATED by scripts/build-seed.py — do not edit by hand.")
    out.append("// Source: workout-tracker/strong-curves.xlsx")
    out.append("// Run `pnpm seed:build` after the source spreadsheet changes.")
    out.append("")
    out.append("import type { Exercise, RoutineTemplate } from '../types';")
    out.append("")

    out.append("// Note: `requiredEquipment` / `instructions` / `diagram` are layered")
    out.append("// on at seed-load time by `src/db/seed-loader.ts` from the enrichment")
    out.append("// map. The generator omits them so it doesn't need to know about")
    out.append("// equipment tagging.")
    out.append("export const STRONG_CURVES_EXERCISES: ReadonlyArray<")
    out.append("  Omit<Exercise, 'isCustom' | 'profileId' | 'requiredEquipment' | 'instructions' | 'diagram'>")
    out.append("> = [")
    for ex in exercises:
        out.append("  {")
        out.append(f"    id: {ts_string(ex.id)},")
        out.append(f"    name: {ts_string(ex.name)},")
        out.append(f"    category: {ts_string(ex.category)},")
        out.append(
            "    primaryMuscles: ["
            + ", ".join(ts_string(m) for m in ex.primary_muscles)
            + "],"
        )
        out.append(
            "    secondaryMuscles: ["
            + ", ".join(ts_string(m) for m in ex.secondary_muscles)
            + "],"
        )
        out.append(f"    measurementType: {ts_string(ex.measurement_type)},")
        out.append(f"    defaultRestSeconds: {ex.default_rest_seconds},")
        out.append(f"    perSide: {'true' if ex.per_side else 'false'},")
        out.append(f"    usesBarbell: {'true' if ex.uses_barbell else 'false'},")
        out.append("  },")
    out.append("] as const;")
    out.append("")

    out.append("type SeedRoutine = Omit<RoutineTemplate, 'profileId' | 'createdAt' | 'updatedAt'>;")
    out.append("")
    out.append("export const STRONG_CURVES_ROUTINES: ReadonlyArray<SeedRoutine> = [")
    for routine in routines:
        out.append("  {")
        out.append(f"    id: {ts_string(routine.id)},")
        out.append(f"    name: {ts_string(routine.name)},")
        out.append(f"    description: {ts_string(routine.description)},")
        out.append("    isSeed: true,")
        out.append("    weeks: [")
        for week in routine.weeks:
            out.append(f"      {{ weekNumber: {week.week_number}, days: [")
            for day in week.days:
                if day.kind == "rest":
                    out.append(
                        f"        {{ dayNumber: {day.day_number}, kind: 'rest', blocks: [] }},"
                    )
                    continue
                label = ts_string(day.workout_label or "")
                out.append(
                    f"        {{ dayNumber: {day.day_number}, kind: 'workout', workoutLabel: {label}, blocks: ["
                )
                for block in day.blocks:
                    out.append(f"          {{ type: {ts_string(block.type)}, exercises: [")
                    for ex in block.exercises:
                        parts = [
                            f"exerciseId: {ts_string(ex.exercise_id)}",
                            f"setCount: {ex.set_count}",
                        ]
                        if ex.reps_min is not None and ex.reps_max is not None:
                            parts.append(
                                f"reps: {{ min: {ex.reps_min}, max: {ex.reps_max} }}"
                            )
                        if ex.duration_min is not None and ex.duration_max is not None:
                            parts.append(
                                f"durationSeconds: {{ min: {ex.duration_min}, max: {ex.duration_max} }}"
                            )
                        if ex.per_side:
                            parts.append("perSide: true")
                        if ex.notes:
                            parts.append(f"notes: {ts_string(ex.notes)}")
                        out.append("            { " + ", ".join(parts) + " },")
                    out.append("          ] },")
                out.append("        ] },")
            out.append("      ] },")
        out.append("    ],")
        out.append("  },")
    out.append("] as const;")
    out.append("")
    return "\n".join(out)


# --- Side-channel for exercise display names -------------------------------

NAME_BY_SLUG: dict[str, str] = {}


def parse_day_capture_names(ws, start_row: int, end_row: int) -> list[SeedBlock]:
    """Same as parse_day, but also populates NAME_BY_SLUG so the codegen
    can render the original capitalisation rather than the slug."""
    blocks: list[SeedBlock] = []
    pending: dict[str, SeedBlock] = {}

    r = start_row + 1
    while r < end_row:
        name_cell = ws.cell(r, WORKOUT_COL).value
        if not isinstance(name_cell, str) or not name_cell.strip():
            r += 1
            continue

        next_row = r + 1
        while next_row < end_row and (
            ws.cell(next_row, WORKOUT_COL).value is None
            or not str(ws.cell(next_row, WORKOUT_COL).value).strip()
        ):
            next_row += 1
        if next_row >= end_row:
            break

        spec = ws.cell(next_row, WORKOUT_COL).value
        if not looks_like_sets_reps(spec):
            break

        clean_name, marker = normalise_exercise_name(name_cell)
        set_count, kwargs = parse_sets_reps(str(spec))

        notes = None
        notes_row = next_row + 1
        if notes_row < end_row:
            notes_value = ws.cell(notes_row, WORKOUT_COL).value
            if isinstance(notes_value, str) and notes_value.strip():
                following = (
                    ws.cell(notes_row + 1, WORKOUT_COL).value
                    if notes_row + 1 < end_row
                    else None
                )
                if not looks_like_sets_reps(following):
                    notes = notes_value.strip()
                    next_row = notes_row

        slug = slugify(clean_name)
        NAME_BY_SLUG.setdefault(slug, clean_name)

        planned = SeedPlannedExercise(
            exercise_id=slug,
            set_count=set_count,
            notes=notes,
            **kwargs,
        )

        if marker:
            letter = marker[0]
            if letter in pending:
                pending[letter].exercises.append(planned)
            else:
                block = SeedBlock(type="superset", exercises=[planned])
                blocks.append(block)
                pending[letter] = block
        else:
            blocks.append(SeedBlock(type="single", exercises=[planned]))
            pending.clear()

        r = next_row + 1

    return blocks


# Override parse_day with the variant that also captures names. (Cleaner than
# threading an output dict through every signature.)
parse_day = parse_day_capture_names  # noqa: F811


# --- Entry point ------------------------------------------------------------

def main() -> int:
    if not XLSX_PATH.exists():
        print(f"error: {XLSX_PATH} not found", file=sys.stderr)
        return 1

    print(f"reading {XLSX_PATH.name}…")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    routines = [
        parse_routine(
            wb,
            name="Strong Curves — Bootyful Beginnings",
            description=(
                "Bret Contreras' beginner glute-focused programme. "
                "Four workouts a week (A/B/A/C) over 8 weeks. Mixes "
                "barbell, dumbbell, and bodyweight work."
            ),
            week_sheet_pattern="Bootyful Beginnings Week {n}",
            week_count=8,
        ),
        parse_routine(
            wb,
            name="Strong Curves — Bodyweight",
            description=(
                "Equipment-free variant of Bootyful Beginnings. Four "
                "workouts a week (A/B/A/C) over 4 weeks."
            ),
            week_sheet_pattern="Bodyweight Week {n}",
            week_count=4,
        ),
    ]

    exercises = collect_exercises(routines)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(ts_emit(routines, exercises), encoding="utf-8")

    total_blocks = sum(
        len(day.blocks)
        for routine in routines
        for week in routine.weeks
        for day in week.days
    )
    print(
        f"wrote {OUTPUT_PATH.relative_to(ROOT)}: "
        f"{len(routines)} routines, {sum(len(r.weeks) for r in routines)} weeks, "
        f"{total_blocks} blocks, {len(exercises)} exercises."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
